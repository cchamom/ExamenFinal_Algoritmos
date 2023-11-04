
import openpyxl
# Crear un nuevo libro de Excel
lista_supermercado= openpyxl.Workbook()

# Seleccionar la hoja activa
hoja = lista_supermercado.active

# Agregar encabezados
hoja['A1'] = "Supermercado"
hoja['B1'] = "Producto"
hoja['C1'] = "Precio"

# Lista de supermercados con nombre de producto y precio
mercado = [
    ("Supermercado A", "Manzanas", 2.99),
    ("Supermercado A", "Peras", 3.49),
    ("Supermercado B", "Naranjas", 2.79),
    ("Supermercado B", "Kiwi", 3.29),
    # Agrega más productos y precios aquí
]

# Agregar datos a la hoja
for row, data in enumerate(mercado, start=2):
    hoja.cell(row=row, column=1, value=data[0])
    hoja.cell(row=row, column=2, value=data[1])
    hoja.cell(row=row, column=3, value=data[2])

# Guardar el archivo Excel
lista_supermercado.save("lista_supermercados.xlsx")

# Cerrar el archivo Excel
lista_supermercado.close()
